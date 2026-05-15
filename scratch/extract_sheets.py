import sys, os, glob
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

# Find the file
files = glob.glob(os.path.join(r'C:\Users\s_ara\OneDrive\Escritorio\Proyectos\agras_comercial', 'Cotizaci*Ecommerce*.xlsm'))
print(f"Found: {files}")
if not files:
    sys.exit(1)

wb = openpyxl.load_workbook(files[0], read_only=True, data_only=True)
print(f"Sheets: {wb.sheetnames}")

ws = wb['VENTA DIRECTA QTC']
print('\n=== VENTA DIRECTA QTC (rows 1-40, cols A-R) ===')
for row in ws.iter_rows(min_row=1, max_row=40, max_col=18, values_only=False):
    for cell in row:
        if cell.value is not None:
            print(f'  {cell.coordinate}: {repr(cell.value)[:120]}')

print('\n=== PRODUCTOS (first 40 rows, cols A-J) ===')
ws2 = wb['PRODUCTOS']
for row in ws2.iter_rows(min_row=1, max_row=40, max_col=10, values_only=False):
    vals = []
    for cell in row:
        if cell.value is not None:
            vals.append(f'{cell.coordinate}={repr(cell.value)[:80]}')
    if vals:
        print('  ' + ' | '.join(vals))

print('\n=== CLIENTES (first 5 rows) ===')
ws3 = wb['CLIENTES']
for row in ws3.iter_rows(min_row=1, max_row=5, max_col=10, values_only=False):
    for cell in row:
        if cell.value is not None:
            print(f'  {cell.coordinate}: {repr(cell.value)[:100]}')

print('\n=== REGISTRO (first 3 rows) ===')
ws4 = wb['REGISTRO']
for row in ws4.iter_rows(min_row=1, max_row=3, max_col=15, values_only=False):
    for cell in row:
        if cell.value is not None:
            print(f'  {cell.coordinate}: {repr(cell.value)[:100]}')

print('\n=== base (first 10 rows) ===')
ws5 = wb['base']
for row in ws5.iter_rows(min_row=1, max_row=10, max_col=10, values_only=False):
    for cell in row:
        if cell.value is not None:
            print(f'  {cell.coordinate}: {repr(cell.value)[:100]}')

wb.close()
