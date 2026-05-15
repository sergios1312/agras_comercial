import sys, glob, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
files = glob.glob(os.path.join(r'C:\Users\s_ara\OneDrive\Escritorio\Proyectos\agras_comercial', 'Cotizaci*Ecommerce*.xlsm'))
wb = openpyxl.load_workbook(files[0], read_only=True, data_only=True)
ws = wb['PRODUCTOS']
kits = set()
for row in ws.iter_rows(min_row=2, max_row=500, max_col=7, values_only=True):
    tipo = str(row[0] or '')
    if 'KIT' in tipo or 'MAVIC' in tipo:
        kits.add(tipo)
        cod_sap = row[2] or ''
        producto = row[3] or ''
        cant = row[4] or ''
        pvp = row[5] or ''
        print(f'{tipo} | {cod_sap} | {producto} | cant={cant} | pvp={pvp}')
print()
print('=== UNIQUE KIT/PRODUCT TYPES ===')
for k in sorted(kits):
    print(f'  {k}')
# Also get all unique TIPO values
all_types = set()
for row in ws.iter_rows(min_row=2, max_row=1000, max_col=1, values_only=True):
    if row[0]:
        all_types.add(str(row[0]))
print()
print('=== ALL UNIQUE TIPOS ===')
for t in sorted(all_types):
    print(f'  {t}')
wb.close()
