import openpyxl
import json
import sys

def extract_excel_structure(file_path):
    try:
        # data_only=False extracts formulas
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        # data_only=True extracts evaluated values
        wb_values = openpyxl.load_workbook(file_path, data_only=True)
        
        output = {}
        
        for sheet_name in wb_formulas.sheetnames:
            sheet_formulas = wb_formulas[sheet_name]
            sheet_values = wb_values[sheet_name]
            
            output[sheet_name] = []
            
            # Limit to reasonable bounds (e.g. 100 rows, 20 columns)
            for row in range(1, 100):
                for col in range(1, 20):
                    cell_f = sheet_formulas.cell(row=row, column=col)
                    cell_v = sheet_values.cell(row=row, column=col)
                    
                    if cell_f.value is not None:
                        # Convert to string to avoid serialization issues
                        val_f = str(cell_f.value)
                        val_v = str(cell_v.value)
                        
                        cell_ref = cell_f.coordinate
                        
                        if val_f.startswith('='):
                            output[sheet_name].append({
                                'cell': cell_ref,
                                'formula': val_f,
                                'value': val_v
                            })
                        else:
                            # only record it if it looks like a label or a number (skip purely empty-looking spaces if any)
                            if val_f.strip() != "":
                                output[sheet_name].append({
                                    'cell': cell_ref,
                                    'text': val_f
                                })
                                
        # Dump as JSON for easy reading
        with open('excel_analysis.json', 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
            
        print("Analysis complete. Saved to excel_analysis.json")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    extract_excel_structure('ROI_AGRAS, GENERADOR; CAMIONETA; MANTENIMIENTOS.xlsx')
